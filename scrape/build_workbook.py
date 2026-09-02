import csv, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

S = "/tmp/claude-0/-home-user-hehe/5430804a-3056-5f03-b302-95e1a915dcc0/scratchpad"
OUT = "/home/user/hehe/out/etsignals_vendor_keywords.xlsx"

kw = list(csv.DictReader(open(f"{S}/all352_clean.csv", encoding="utf-8")))
summ = list(csv.DictReader(open(f"{S}/summary.csv", encoding="utf-8")))
site = {s["domain"]: s["url"] for s in summ}

TYPE = {"product": "Product", "solution": "Solution", "use_case": "Use case",
        "platform": "Platform", "technology": "Technology", "offering": "Offering",
        "capability": "Capability", "service": "Service", "industry": "Industry",
        "nav": "Navigation"}
FOUND = {"sitemap": "Sitemap", "nav_header": "Header menu", "nav_footer": "Footer menu"}
RANK = {"product": 0, "solution": 1, "use_case": 1, "platform": 2, "technology": 2,
        "offering": 2, "capability": 3, "service": 3, "nav": 4, "industry": 5}

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BODY = Font(name=ARIAL, size=10)
LINK = Font(name=ARIAL, size=10, color="0563C1", underline="single")

wb = Workbook()

# ---------------------------------------------------------------- Read me
rm = wb.active
rm.title = "Read me"
lines = [
    ("ETSignals — vendor keyword extraction", True),
    ("", False),
    ("What this is", True),
    ("Every product, solution, use case and technology term found on each sponsor's own", False),
    ("website, one keyword per row, with the page it came from.", False),
    ("", False),
    ("Sheets", True),
    ("Keywords        one row per keyword, with a link to the page it was found on", False),
    ("Vendors         one row per sponsor: what was fetched and how much was found", False),
    ("Not extracted   the 56 sponsors that returned nothing, and why", False),
    ("", False),
    ("How the keywords were found", True),
    ("1. sitemap.xml — URL slugs under /products/, /solutions/, /use-cases/ and similar", False),
    ("2. the header menu — where Products, Solutions and Platform normally live", False),
    ("3. the footer menu — often a flatter, more complete product list", False),
    ("", False),
    ("Checking a keyword", True),
    ("The 'Source URL' column links to the live page the keyword was read from.", False),
    ("The 'Vendor file' column links to that vendor's folder in vendors_out, which also", False),
    ("holds rejected.csv, sitemap_urls.txt and fetch_log.txt.", False),
    ("For those links to open, unzip vendors_out.zip into the SAME FOLDER as this workbook.", False),
    ("", False),
    ("Counts", True),
    ("'Keywords found' records what this extraction retrieved on 2 September 2026.", False),
    ("It is a fact about the run, not a live formula, so it will not change if rows are", False),
    ("filtered or edited. Re-run the extractor to refresh it.", False),
    ("", False),
    ("Not yet done", True),
    ("Keywords are as the vendor writes them. They have not yet been mapped to shared", False),
    ("concepts (so 'aiSIEM' and 'Percept XDR' do not yet resolve to SIEM and XDR), and", False),
    ("they have not been assigned to the ETSignals category taxonomy.", False),
]
for i, (text, bold) in enumerate(lines, start=1):
    c = rm.cell(row=i, column=1, value=text)
    c.font = Font(name=ARIAL, size=11, bold=bold)
rm.column_dimensions["A"].width = 95
rm.sheet_view.showGridLines = False

# ---------------------------------------------------------------- Keywords
ws = wb.create_sheet("Keywords")
headers = ["Vendor", "Vendor site", "Keyword", "Type", "Found in", "Source URL", "Vendor file"]
ws.append(headers)
kw.sort(key=lambda r: (r["domain"], RANK.get(r["bucket"], 4), r["label"].lower()))
for r in kw:
    d = r["domain"]
    ws.append([d, site.get(d, f"https://{d}"), r["label"], TYPE.get(r["bucket"], r["bucket"]),
               FOUND.get(r["source"], r["source"]), r["url"], f"vendors_out/{d}/keywords.csv"])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row:
        c.font = BODY
    row[1].hyperlink, row[1].font = row[1].value, LINK
    row[5].hyperlink, row[5].font = row[5].value, LINK
    row[6].hyperlink, row[6].font = row[6].value, LINK

# ---------------------------------------------------------------- Vendors
vs = wb.create_sheet("Vendors")
vs.append(["Vendor", "Site", "Status", "Keywords found", "Sitemap URLs seen",
           "Pages fetched", "Vendor file", "Note"])
kw_count = collections.Counter(r["domain"] for r in kw)
summ.sort(key=lambda s: s["domain"])
for i, s in enumerate(summ, start=2):
    status = "OK" if s["homepage"] == "ok" and int(s["keywords_kept"] or 0) else "No keywords"
    note = s["error"][:120] if s["error"] else ""
    vs.append([s["domain"], s["url"], status,
               kw_count[s["domain"]],
               int(s["sitemap_urls_seen"] or 0), int(s["pages_fetched"] or 0),
               f"vendors_out/{s['domain']}/", note])
for row in vs.iter_rows(min_row=2, max_row=vs.max_row):
    for c in row:
        c.font = BODY
    row[1].hyperlink, row[1].font = row[1].value, LINK
    row[6].hyperlink, row[6].font = row[6].value, LINK

# ---------------------------------------------------------------- Failures
fs = wb.create_sheet("Not extracted")
fs.append(["Vendor", "Site", "Reason", "Category"])
for s in summ:
    if not s["error"]:
        continue
    e = s["error"]
    cat = ("Bot wall" if "HTTP" in e else
           "JavaScript-rendered menu" if "JS-rendered" in e else "Network or certificate")
    fs.append([s["domain"], s["url"], e[:150], cat])
for row in fs.iter_rows(min_row=2, max_row=fs.max_row):
    for c in row:
        c.font = BODY
    row[1].hyperlink, row[1].font = row[1].value, LINK

# ---------------------------------------------------------------- formatting
WIDTHS = {"Keywords": [24, 34, 46, 13, 13, 62, 40],
          "Vendors": [26, 36, 14, 15, 18, 14, 30, 60],
          "Not extracted": [26, 36, 90, 26]}
for name, widths in WIDTHS.items():
    sh = wb[name]
    for j, w in enumerate(widths, start=1):
        sh.column_dimensions[get_column_letter(j)].width = w
    for c in sh[1]:
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(vertical="center")
    sh.row_dimensions[1].height = 22
    sh.freeze_panes = "A2"
    sh.auto_filter.ref = f"A1:{get_column_letter(sh.max_column)}{sh.max_row}"

wb.save(OUT)
print("written:", OUT, f"({len(kw):,} keyword rows, {len(summ)} vendors)")
